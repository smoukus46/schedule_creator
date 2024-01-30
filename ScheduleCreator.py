import sys 
from PyQt6 import QtWidgets, QtGui, QtCore
import ctypes
import design
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

class ExampleApp(QtWidgets.QMainWindow, design.Ui_MainWindow):
    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        # Инициализация данных
        self.load_workout_data()
        self.load_trainer_data()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна

        for item_workout_text in self.data_workout:
            item = QtWidgets.QListWidgetItem(item_workout_text)
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.workout_list.addItem(item)
        for item_trainer_text in self.data_trainer:
            item = QtWidgets.QListWidgetItem(item_trainer_text)
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.trainer_list.addItem(item)    

    def remove_trainer_btn_was_clicked(self):
        item = self.trainer_list.currentItem()
        
        if item is not None:
            row = self.trainer_list.row(item)
            self.trainer_list.takeItem(row)
            self.data_trainer.remove(item.text())
            self.save_trainer_data()

    def remove_workout_btn_was_clicked(self):
        item = self.workout_list.currentItem()
        
        if item is not None:
            row = self.workout_list.row(item)
            self.workout_list.takeItem(row)
            self.data_workout.remove(item.text())
            self.save_workout_data()
    
    def add_trainer_btn_was_clicked(self):
        text = self.trainer_list_line_edit.text()
        
        if text:
            item = QtWidgets.QListWidgetItem(text)
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.trainer_list.addItem(item)
            self.data_trainer.append(text)
            self.save_trainer_data()

            # Очищаем QLineEdit
            self.trainer_list_line_edit.clear()

    def add_workout_btn_was_clicked(self):
        text = self.workout_list_line_edit.text()
        
        if text:
            item = QtWidgets.QListWidgetItem(text)
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.workout_list.addItem(item)           
            self.data_workout.append(text)
            self.save_workout_data()

            # Очищаем QLineEdit
            self.workout_list_line_edit.clear()

    def load_workout_data(self):
        try:
            with open("data_workout.txt", "r") as file:
                self.data_workout = file.read().splitlines()
        except FileNotFoundError:
            self.data_workout = []

    def save_workout_data(self):
        with open("data_workout.txt", "w") as file:
            file.write("\n".join(self.data_workout))

    def load_trainer_data(self):
        try:
            with open("data_trainer.txt", "r") as file:
                self.data_trainer = file.read().splitlines()
        except FileNotFoundError:
            self.data_trainer = []

    def save_trainer_data(self):
        with open("data_trainer.txt", "w") as file:
            file.write("\n".join(self.data_trainer))

    def check_inputs(self):
            # Проверяем, заполнены ли оба LineEdit
            if self.month_edit.text() and self.gym_edit.text():
                self.done_btn.setEnabled(True)
            else:
                self.done_btn.setEnabled(False)
            
    def open_modal_dialog(self):
        modal_dialog = design.ModalDialog(self)
        modal_dialog.exec()

    def open_except_modal_dialog(self):
        except_modal_dialog = design.ExceptModalDialog(self)
        except_modal_dialog.exec()

    # Путь к рабочему столу на Windows
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

    # Относительный путь к файлу на рабочем столе
    relative_path = os.path.join(desktop_path, 'Расписание_тренировок_2024.xlsx')

    wb = load_workbook(relative_path)

    def cell_coloring(self, cell_text, cell_coordinate):
        text = self.month_edit.text().capitalize()
        sheet = self.wb[str(text)]
        if 'Лера' in cell_text:
           sheet[cell_coordinate].fill = PatternFill('solid', fgColor="D2691E")
        elif 'Ксюша' in cell_text:
            sheet[cell_coordinate].fill = PatternFill('solid', fgColor="FFDEAD")
        elif 'Яна' in cell_text:
            sheet[cell_coordinate].fill = PatternFill('solid', fgColor="228B22")
        elif 'Таня' in cell_text:
            sheet[cell_coordinate].fill = PatternFill('solid', fgColor="4169E1")
        elif 'Юля' in cell_text:
            sheet[cell_coordinate].fill = PatternFill('solid', fgColor="B0C4DE")
        elif 'Лена' in cell_text:
            sheet[cell_coordinate].fill = PatternFill('solid', fgColor="EE82EE")
        elif 'Настя' in cell_text:
            sheet[cell_coordinate].fill = PatternFill('solid', fgColor="9ACD32")
        elif 'Настя Е.' in cell_text:
            sheet[cell_coordinate].fill = PatternFill('solid', fgColor="DAA520")
        self.wb.save(self.relative_path)
        self.wb.close()


    def write_cells(self):
        try: 
            text = self.month_edit.text().capitalize()
            gym = self.gym_edit.text().capitalize()
            sheet = self.wb[str(text)]
            
            if gym == 'Зал 1':
                
                # Заполняем понедельник
                sheet['B4'] = self.monday_10_textEdit.toPlainText()
                self.cell_coloring(self.monday_10_textEdit.toPlainText(), 'B4')
                sheet['B5'] = self.monday_11_textEdit.toPlainText()
                self.cell_coloring(self.monday_11_textEdit.toPlainText(), 'B5')
                sheet['B6'] = self.monday_12_textEdit.toPlainText()
                self.cell_coloring(self.monday_12_textEdit.toPlainText(), 'B6')
                sheet['B7'] = self.monday_13_textEdit.toPlainText()
                self.cell_coloring(self.monday_13_textEdit.toPlainText(), 'B7')
                sheet['B9'] = self.monday_16_textEdit.toPlainText()
                self.cell_coloring(self.monday_16_textEdit.toPlainText(), 'B9')
                sheet['B10'] = self.monday_17_textEdit.toPlainText()
                self.cell_coloring(self.monday_17_textEdit.toPlainText(), 'B10')
                sheet['B11'] = self.monday_18_textEdit.toPlainText()
                self.cell_coloring(self.monday_18_textEdit.toPlainText(), 'B11')
                sheet['B12'] = self.monday_19_textEdit.toPlainText()
                self.cell_coloring(self.monday_19_textEdit.toPlainText(), 'B12')
                sheet['B13'] = self.monday_20_textEdit.toPlainText()
                self.cell_coloring(self.monday_20_textEdit.toPlainText(), 'B13')

                # Заполняем вторник
                sheet['C4'] = self.tuesday_10_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_10_textEdit.toPlainText(), 'C4')
                sheet['C5'] = self.tuesday_11_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_11_textEdit.toPlainText(), 'C5')
                sheet['C6'] = self.tuesday_12_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_12_textEdit.toPlainText(), 'C6')
                sheet['C7'] = self.tuesday_13_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_13_textEdit.toPlainText(), 'C7')
                sheet['C9'] = self.tuesday_16_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_16_textEdit.toPlainText(), 'C9')
                sheet['C10'] = self.tuesday_17_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_17_textEdit.toPlainText(), 'C10')
                sheet['C11'] = self.tuesday_18_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_18_textEdit.toPlainText(), 'C11')
                sheet['C12'] = self.tuesday_19_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_19_textEdit.toPlainText(), 'C12')
                sheet['C13'] = self.tuesday_20_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_20_textEdit.toPlainText(), 'C13')

                # Заполняем среду
                sheet['D4'] = self.wednesday_10_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_10_textEdit.toPlainText(), 'D4')
                sheet['D5'] = self.wednesday_11_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_11_textEdit.toPlainText(), 'D5')
                sheet['D6'] = self.wednesday_12_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_12_textEdit.toPlainText(), 'D6')
                sheet['D7'] = self.wednesday_13_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_13_textEdit.toPlainText(), 'D7')
                sheet['D9'] = self.wednesday_16_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_16_textEdit.toPlainText(), 'D9')
                sheet['D10'] = self.wednesday_17_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_17_textEdit.toPlainText(), 'D10')
                sheet['D11'] = self.wednesday_18_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_18_textEdit.toPlainText(), 'D11')
                sheet['D12'] = self.wednesday_19_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_19_textEdit.toPlainText(), 'D12')
                sheet['D13'] = self.wednesday_20_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_20_textEdit.toPlainText(), 'D13')

                # Заполняем четверг
                sheet['E4'] = self.thursday_10_textEdit.toPlainText()
                self.cell_coloring(self.thursday_10_textEdit.toPlainText(), 'E4')
                sheet['E5'] = self.thursday_11_textEdit.toPlainText()
                self.cell_coloring(self.thursday_11_textEdit.toPlainText(), 'E5')
                sheet['E6'] = self.thursday_12_textEdit.toPlainText()
                self.cell_coloring(self.thursday_12_textEdit.toPlainText(), 'E6')
                sheet['E7'] = self.thursday_13_textEdit.toPlainText()
                self.cell_coloring(self.thursday_13_textEdit.toPlainText(), 'E7')
                sheet['E9'] = self.thursday_16_textEdit.toPlainText()
                self.cell_coloring(self.thursday_16_textEdit.toPlainText(), 'E9')
                sheet['E10'] = self.thursday_17_textEdit.toPlainText()
                self.cell_coloring(self.thursday_17_textEdit.toPlainText(), 'E10')
                sheet['E11'] = self.thursday_18_textEdit.toPlainText()
                self.cell_coloring(self.thursday_18_textEdit.toPlainText(), 'E11')
                sheet['E12'] = self.thursday_19_textEdit.toPlainText()
                self.cell_coloring(self.thursday_19_textEdit.toPlainText(), 'E12')
                sheet['E13'] = self.thursday_20_textEdit.toPlainText()
                self.cell_coloring(self.thursday_20_textEdit.toPlainText(), 'E13')

                # Заполняем пятницу
                sheet['F4'] = self.friday_10_textEdit.toPlainText()
                self.cell_coloring(self.friday_10_textEdit.toPlainText(), 'F4')
                sheet['F5'] = self.friday_11_textEdit.toPlainText()
                self.cell_coloring(self.friday_11_textEdit.toPlainText(), 'F5')
                sheet['F6'] = self.friday_12_textEdit.toPlainText()
                self.cell_coloring(self.friday_12_textEdit.toPlainText(), 'F6')
                sheet['F7'] = self.friday_13_textEdit.toPlainText()
                self.cell_coloring(self.friday_13_textEdit.toPlainText(), 'F7')
                sheet['F9'] = self.friday_16_textEdit.toPlainText()
                self.cell_coloring(self.friday_16_textEdit.toPlainText(), 'F9')
                sheet['F10'] = self.friday_17_textEdit.toPlainText()
                self.cell_coloring(self.friday_17_textEdit.toPlainText(), 'F10')
                sheet['F11'] = self.friday_18_textEdit.toPlainText()
                self.cell_coloring(self.friday_18_textEdit.toPlainText(), 'F11')
                sheet['F12'] = self.friday_19_textEdit.toPlainText()
                self.cell_coloring(self.friday_19_textEdit.toPlainText(), 'F12')
                sheet['F13'] = self.friday_20_textEdit.toPlainText()
                self.cell_coloring(self.friday_20_textEdit.toPlainText(), 'F13')

                # Заполняем субботу
                sheet['G4'] = self.saturday_10_textEdit.toPlainText()
                self.cell_coloring(self.saturday_10_textEdit.toPlainText(), 'G4')
                sheet['G5'] = self.saturday_11_textEdit.toPlainText()
                self.cell_coloring(self.saturday_11_textEdit.toPlainText(), 'G5')
                sheet['G6'] = self.saturday_12_textEdit.toPlainText()
                self.cell_coloring(self.saturday_12_textEdit.toPlainText(), 'G6')
                sheet['G7'] = self.saturday_13_textEdit.toPlainText()
                self.cell_coloring(self.saturday_13_textEdit.toPlainText(), 'G7')
                sheet['G9'] = self.saturday_16_textEdit.toPlainText()
                self.cell_coloring(self.saturday_16_textEdit.toPlainText(), 'G9')
                sheet['G10'] = self.saturday_17_textEdit.toPlainText()
                self.cell_coloring(self.saturday_17_textEdit.toPlainText(), 'G10')
                sheet['G11'] = self.saturday_18_textEdit.toPlainText()
                self.cell_coloring(self.saturday_18_textEdit.toPlainText(), 'G11')
                sheet['G12'] = self.saturday_19_textEdit.toPlainText()
                self.cell_coloring(self.saturday_19_textEdit.toPlainText(), 'G12')
                sheet['G13'] = self.saturday_20_textEdit.toPlainText()
                self.cell_coloring(self.saturday_20_textEdit.toPlainText(), 'G13')

                # Заполняем воскресенье
                sheet['H4'] = self.sunday_10_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H4')
                sheet['H5'] = self.sunday_11_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H5')
                sheet['H6'] = self.sunday_12_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H6')
                sheet['H7'] = self.sunday_13_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H7')
                sheet['H9'] = self.sunday_16_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H9')
                sheet['H10'] = self.sunday_17_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H10')
                sheet['H11'] = self.sunday_18_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H11')
                sheet['H12'] = self.sunday_19_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H12')
                sheet['H13'] = self.sunday_20_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H13')

                self.wb.save(self.relative_path)
                self.wb.close()
                self.open_modal_dialog()
            
            elif gym == 'Зал 2':
                
                # Заполняем понедельник
                sheet['B18'] = self.monday_10_textEdit.toPlainText()
                self.cell_coloring(self.monday_10_textEdit.toPlainText(), 'B18')
                sheet['B19'] = self.monday_11_textEdit.toPlainText()
                self.cell_coloring(self.monday_11_textEdit.toPlainText(), 'B19')
                sheet['B20'] = self.monday_12_textEdit.toPlainText()
                self.cell_coloring(self.monday_12_textEdit.toPlainText(), 'B20')
                sheet['B21'] = self.monday_13_textEdit.toPlainText()
                self.cell_coloring(self.monday_13_textEdit.toPlainText(), 'B21')
                sheet['B23'] = self.monday_16_textEdit.toPlainText()
                self.cell_coloring(self.monday_16_textEdit.toPlainText(), 'B23')
                sheet['B24'] = self.monday_17_textEdit.toPlainText()
                self.cell_coloring(self.monday_17_textEdit.toPlainText(), 'B24')
                sheet['B25'] = self.monday_18_textEdit.toPlainText()
                self.cell_coloring(self.monday_18_textEdit.toPlainText(), 'B25')
                sheet['B26'] = self.monday_19_textEdit.toPlainText()
                self.cell_coloring(self.monday_19_textEdit.toPlainText(), 'B26')
                sheet['B27'] = self.monday_20_textEdit.toPlainText()
                self.cell_coloring(self.monday_20_textEdit.toPlainText(), 'B27')

                # Заполняем вторник
                sheet['C18'] = self.tuesday_10_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_10_textEdit.toPlainText(), 'C18')
                sheet['C19'] = self.tuesday_11_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_11_textEdit.toPlainText(), 'C19')
                sheet['C20'] = self.tuesday_12_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_12_textEdit.toPlainText(), 'C20')
                sheet['C21'] = self.tuesday_13_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_13_textEdit.toPlainText(), 'C21')
                sheet['C23'] = self.tuesday_16_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_16_textEdit.toPlainText(), 'C23')
                sheet['C24'] = self.tuesday_17_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_17_textEdit.toPlainText(), 'C24')
                sheet['C25'] = self.tuesday_18_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_18_textEdit.toPlainText(), 'C25')
                sheet['C26'] = self.tuesday_19_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_19_textEdit.toPlainText(), 'C26')
                sheet['C27'] = self.tuesday_20_textEdit.toPlainText()
                self.cell_coloring(self.tuesday_20_textEdit.toPlainText(), 'C27')

                # Заполняем среду
                sheet['D18'] = self.wednesday_10_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_10_textEdit.toPlainText(), 'D18')
                sheet['D19'] = self.wednesday_11_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_11_textEdit.toPlainText(), 'D19')
                sheet['D20'] = self.wednesday_12_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_12_textEdit.toPlainText(), 'D20')
                sheet['D21'] = self.wednesday_13_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_13_textEdit.toPlainText(), 'D21')
                sheet['D23'] = self.wednesday_16_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_16_textEdit.toPlainText(), 'D23')
                sheet['D24'] = self.wednesday_17_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_17_textEdit.toPlainText(), 'D24')
                sheet['D25'] = self.wednesday_18_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_18_textEdit.toPlainText(), 'D25')
                sheet['D26'] = self.wednesday_19_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_19_textEdit.toPlainText(), 'D26')
                sheet['D27'] = self.wednesday_20_textEdit.toPlainText()
                self.cell_coloring(self.wednesday_20_textEdit.toPlainText(), 'D27')
            
                # Заполняем четверг
                sheet['E18'] = self.thursday_10_textEdit.toPlainText()
                self.cell_coloring(self.thursday_10_textEdit.toPlainText(), 'E18')
                sheet['E19'] = self.thursday_11_textEdit.toPlainText()
                self.cell_coloring(self.thursday_11_textEdit.toPlainText(), 'E19')
                sheet['E20'] = self.thursday_12_textEdit.toPlainText()
                self.cell_coloring(self.thursday_12_textEdit.toPlainText(), 'E20')
                sheet['E21'] = self.thursday_13_textEdit.toPlainText()
                self.cell_coloring(self.thursday_13_textEdit.toPlainText(), 'E21')
                sheet['E23'] = self.thursday_16_textEdit.toPlainText()
                self.cell_coloring(self.thursday_16_textEdit.toPlainText(), 'E23')
                sheet['E24'] = self.thursday_17_textEdit.toPlainText()
                self.cell_coloring(self.thursday_17_textEdit.toPlainText(), 'E24')
                sheet['E25'] = self.thursday_18_textEdit.toPlainText()
                self.cell_coloring(self.thursday_18_textEdit.toPlainText(), 'E25')
                sheet['E26'] = self.thursday_19_textEdit.toPlainText()
                self.cell_coloring(self.thursday_19_textEdit.toPlainText(), 'E26')
                sheet['E27'] = self.thursday_20_textEdit.toPlainText()
                self.cell_coloring(self.thursday_20_textEdit.toPlainText(), 'E27')
                
                # Заполняем пятницу
                sheet['F18'] = self.friday_10_textEdit.toPlainText()
                self.cell_coloring(self.friday_10_textEdit.toPlainText(), 'F18')
                sheet['F19'] = self.friday_11_textEdit.toPlainText()
                self.cell_coloring(self.friday_11_textEdit.toPlainText(), 'F19')
                sheet['F20'] = self.friday_12_textEdit.toPlainText()
                self.cell_coloring(self.friday_12_textEdit.toPlainText(), 'F20')
                sheet['F21'] = self.friday_13_textEdit.toPlainText()
                self.cell_coloring(self.friday_13_textEdit.toPlainText(), 'F21')
                sheet['F23'] = self.friday_16_textEdit.toPlainText()
                self.cell_coloring(self.friday_16_textEdit.toPlainText(), 'F23')
                sheet['F24'] = self.friday_17_textEdit.toPlainText()
                self.cell_coloring(self.friday_17_textEdit.toPlainText(), 'F24')
                sheet['F25'] = self.friday_18_textEdit.toPlainText()
                self.cell_coloring(self.friday_18_textEdit.toPlainText(), 'F25')
                sheet['F26'] = self.friday_19_textEdit.toPlainText()
                self.cell_coloring(self.friday_19_textEdit.toPlainText(), 'F26')
                sheet['F27'] = self.friday_20_textEdit.toPlainText()
                self.cell_coloring(self.friday_20_textEdit.toPlainText(), 'F27')

                # Заполняем субботу
                sheet['G18'] = self.saturday_10_textEdit.toPlainText()
                self.cell_coloring(self.saturday_10_textEdit.toPlainText(), 'G18')
                sheet['G19'] = self.saturday_11_textEdit.toPlainText()
                self.cell_coloring(self.saturday_11_textEdit.toPlainText(), 'G19')
                sheet['G20'] = self.saturday_12_textEdit.toPlainText()
                self.cell_coloring(self.saturday_12_textEdit.toPlainText(), 'G20')
                sheet['G21'] = self.saturday_13_textEdit.toPlainText()
                self.cell_coloring(self.saturday_13_textEdit.toPlainText(), 'G21')
                sheet['G23'] = self.saturday_16_textEdit.toPlainText()
                self.cell_coloring(self.saturday_16_textEdit.toPlainText(), 'G23')
                sheet['G24'] = self.saturday_17_textEdit.toPlainText()
                self.cell_coloring(self.saturday_17_textEdit.toPlainText(), 'G24')
                sheet['G25'] = self.saturday_18_textEdit.toPlainText()
                self.cell_coloring(self.saturday_18_textEdit.toPlainText(), 'G25')
                sheet['G26'] = self.saturday_19_textEdit.toPlainText()
                self.cell_coloring(self.saturday_19_textEdit.toPlainText(), 'G26')
                sheet['G27'] = self.saturday_20_textEdit.toPlainText()
                self.cell_coloring(self.saturday_20_textEdit.toPlainText(), 'G27')

                # Заполняем воскресенье
                sheet['H18'] = self.sunday_10_textEdit.toPlainText()
                self.cell_coloring(self.sunday_10_textEdit.toPlainText(), 'H18')
                sheet['H19'] = self.sunday_11_textEdit.toPlainText()
                self.cell_coloring(self.sunday_11_textEdit.toPlainText(), 'H19')
                sheet['H20'] = self.sunday_12_textEdit.toPlainText()
                self.cell_coloring(self.sunday_12_textEdit.toPlainText(), 'H20')
                sheet['H21'] = self.sunday_13_textEdit.toPlainText()
                self.cell_coloring(self.sunday_13_textEdit.toPlainText(), 'H21')
                sheet['H23'] = self.sunday_16_textEdit.toPlainText()
                self.cell_coloring(self.sunday_16_textEdit.toPlainText(), 'H23')
                sheet['H24'] = self.sunday_17_textEdit.toPlainText()
                self.cell_coloring(self.sunday_17_textEdit.toPlainText(), 'H24')
                sheet['H25'] = self.sunday_18_textEdit.toPlainText()
                self.cell_coloring(self.sunday_18_textEdit.toPlainText(), 'H25')
                sheet['H26'] = self.sunday_19_textEdit.toPlainText()
                self.cell_coloring(self.sunday_19_textEdit.toPlainText(), 'H26')
                sheet['H27'] = self.sunday_20_textEdit.toPlainText()
                self.cell_coloring(self.sunday_20_textEdit.toPlainText(), 'H27')

                self.wb.save(self.relative_path)
                self.wb.close()
                self.open_modal_dialog()

            else:
                self.wb.save(self.relative_path)
                self.wb.close()
                self.open_except_modal_dialog()

            
        except:
            self.wb.save(self.relative_path)
            self.wb.close()
            self.open_except_modal_dialog()
        
if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    myappid = 'mycompany.myproduct.subproduct.version'
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    app.setWindowIcon(QtGui.QIcon('window_icon3.svg'))
    app.setStyle("Fusion")
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.setWindowIcon(QtGui.QIcon('window_icon3.svg'))
    window.show()  # Показываем окно
    sys.exit(app.exec())  # и запускаем приложение